"""
run_einleser.py
---------------
Liest den Reiter 'Einleser' aus Bachert_Standard_Einleser_AUTOMATE.xlsm,
mappt alle einzigartigen Konten per KI direkt auf Dummy-Pool-Positionen
und erzeugt die Lucanet-Import-Datei mit den Dummy-OIDs.

Flow:
  1. Konten aus 'Einleser'-Sheet extrahieren (col A='Konto', col C=Name)
  2. Dummy-Pool laden (Dummykonten_Zuordnung_hart.xlsx, Reiter 'Mapping')
  3. Pool-Überpositionen als KI-Whitelist verwenden
  4. KI (gpt-5-mini) mapped jedes Konto auf eine dieser Pool-Positionen (parallel)
  5. Ersten freien Dummy-OID pro Position zuweisen
  6. Lucanet-Export schreiben

Ausgaben (./output/einleser/):
    Mapping_Einleser.xlsx         — KI-Ergebnis (Konto → Pool-Position)
    Lucanet_Mapping_Einleser.xlsx — Lucanet-Importformat mit Dummy-OIDs

Verwendung:
    python tools/run_einleser.py
    python tools/run_einleser.py --einleser Examples/Bachert_Standard_Einleser_AUTOMATE.xlsm
"""

import os
import sys
import logging
import argparse

import openpyxl
import pandas as pd
from pathlib import Path

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from src.llm_client import LLMClient
from src.mapping import map_accounts
from src.dummy_mapper import (
    load_dummy_pool,
    pool_positions_as_targets,
    assign_dummy_ids,
    build_lucanet_df,
    save_lucanet_xlsx,
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
log = logging.getLogger("run_einleser")


# ---------------------------------------------------------------------------
# Konten aus 'Einleser'-Sheet laden
# ---------------------------------------------------------------------------

def load_accounts_from_einleser(einleser_path: str | Path) -> pd.DataFrame:
    """
    Reiter 'Einleser':
        Col A: 'Konto'                ← Zeilentyp-Marker
        Col C: Konto-Nr. + Kontoname  ← z.B. "1000 Kasse"
        Col F-H: Beträge (je nach Unternehmensanzahl / Perioden)

    Gibt deduplizierten DataFrame mit row_type='ACCOUNT', konto_nr, konto_name zurück.
    """
    wb = openpyxl.load_workbook(
        str(einleser_path), read_only=True, data_only=True, keep_vba=False
    )
    if "Einleser" not in wb.sheetnames:
        raise ValueError(f"Kein Reiter 'Einleser' in {einleser_path}. Vorhanden: {wb.sheetnames}")

    ws = wb["Einleser"]
    seen: set[str] = set()
    rows_out = []

    for row in ws.iter_rows(values_only=True):
        if not row[0] or str(row[0]).strip() != "Konto":
            continue
        cell_c = row[2]
        if not cell_c:
            continue

        val = str(cell_c).strip()
        if not val or val in seen:
            continue
        seen.add(val)

        parts = val.split(" ", 1)
        try:
            konto_nr   = int(parts[0])
            konto_name = parts[1].strip() if len(parts) > 1 else parts[0]
        except (ValueError, IndexError):
            konto_nr   = None
            konto_name = val

        rows_out.append({
            "row_type":          "ACCOUNT",
            "konto_nr":          konto_nr,
            "konto_name":        konto_name,
            "amount_normalized": None,
        })

    wb.close()
    df = pd.DataFrame(rows_out)
    log.info("Einzigartige Konten aus 'Einleser': %d", len(df))
    return df


# ---------------------------------------------------------------------------
# Hauptlauf
# ---------------------------------------------------------------------------

def run_einleser(
    einleser_path:   str = "Examples/Bachert_Standard_Einleser_AUTOMATE.xlsm",
    dummy_pool_path: str = "Examples/Dummykonten_Zuordnung_hart.xlsx",
    out_dir_str:     str = "./output/einleser",
) -> None:
    out_dir = Path(out_dir_str)
    out_dir.mkdir(parents=True, exist_ok=True)

    llm = LLMClient(
        model="gpt-5-mini-2025-08-07",
        cache_db_path=out_dir / "llm_cache.db",
    )

    # 1. Dummy-Pool laden
    log.info("Lade Dummy-Pool aus %s", dummy_pool_path)
    dummy_pool = load_dummy_pool(dummy_pool_path)

    # 2. Pool-Positionen als KI-Whitelist
    targets = pool_positions_as_targets(dummy_pool)
    log.info("  → %d Pool-Positionen als Whitelist", len(targets))

    # 3. Konten aus Einleser lesen
    log.info("Lese Konten aus %s", einleser_path)
    accounts_df = load_accounts_from_einleser(einleser_path)
    account_count = len(accounts_df)
    log.info("  → %d einzigartige Konten", account_count)

    # 4. KI-Mapping — alle Batches parallel (max 10 Worker)
    log.info("KI-Mapping startet (%d Konten, batch_size=50, parallel) ...", account_count)
    mapped_df = map_accounts(llm, accounts_df, targets, batch_size=50, reasoning_effort="medium")

    # 5. KI-Mapping-Excel speichern
    account_rows = mapped_df[mapped_df["row_type"] == "ACCOUNT"].copy()
    ki_df = pd.DataFrame({
        "Konto-Nr":                   account_rows["konto_nr"].values,
        "Konto-Name":                 account_rows["konto_name"].values,
        "Kontonr & Bezeichnung":      account_rows.apply(
            lambda r: f"{str(r.get('konto_nr','') or '').strip()} {str(r.get('konto_name','') or '').strip()}",
            axis=1,
        ).values,
        "Pool-Position (KI-Ergebnis)": account_rows["target_overpos_name"].values,
        "Pool-Key":                    account_rows["target_overpos_id"].values,
        "Confidence":                  account_rows["confidence"].values,
        "Begründung":                  account_rows["rationale_short"].values,
    })

    ki_path = out_dir / "Mapping_Einleser.xlsx"
    ki_df.to_excel(ki_path, index=False, sheet_name="Mapping Einleser")
    log.info("KI-Mapping gespeichert: %s", ki_path)

    # 6. Dummy-OID zuweisen → Lucanet-Export
    log.info("Weise Dummy-OIDs zu ...")
    mapped_with_dummies = assign_dummy_ids(mapped_df, dummy_pool)
    lucanet_df = build_lucanet_df(mapped_with_dummies)
    lucanet_path = out_dir / "Lucanet_Mapping_Einleser.xlsx"
    save_lucanet_xlsx(lucanet_df, lucanet_path)

    # Statistik
    mapped_ok      = len(lucanet_df)
    without_dummy  = account_count - mapped_ok
    log.info("Statistik: %d total | %d mit OID | %d ohne", account_count, mapped_ok, without_dummy)

    if without_dummy > 0:
        no_dummy = mapped_with_dummies[
            (mapped_with_dummies["row_type"] == "ACCOUNT") &
            (mapped_with_dummies["dummy_oid"].isna() | (mapped_with_dummies["dummy_oid"] == ""))
        ]
        for _, r in no_dummy.iterrows():
            log.warning("  ⚠ Kein OID: '%s %s' → '%s'",
                        r.get("konto_nr",""), r.get("konto_name",""), r.get("target_overpos_name",""))

    print(f"\nDONE! {account_count} Konten verarbeitet.")
    print(f"  KI-Mapping:      {ki_path}")
    print(f"  Lucanet-Export:  {lucanet_path}  ({mapped_ok} Zeilen mit Dummy-OID)")
    if without_dummy:
        print(f"  ⚠  Ohne Dummy-OID: {without_dummy} (UNMAPPED oder Pool erschöpft — s. Log)")

    llm.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Bachert Einleser → Lucanet Dummy-Mapping")
    parser.add_argument("--einleser",   default="Examples/Bachert_Standard_Einleser_AUTOMATE.xlsm")
    parser.add_argument("--dummy-pool", default="Examples/Dummykonten_Zuordnung_hart.xlsx")
    parser.add_argument("--out",        default="./output/einleser")
    args = parser.parse_args()

    run_einleser(
        einleser_path=args.einleser,
        dummy_pool_path=args.dummy_pool,
        out_dir_str=args.out,
    )
