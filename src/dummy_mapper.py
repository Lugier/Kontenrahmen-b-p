"""
dummy_mapper.py
---------------
Zwei Aufgaben:

1. pool_positions_as_targets()
   Wandelt die Überpositionen aus dem Dummy-Pool (Dummykonten_Zuordnung_hart.xlsx)
   in TargetPosition-Objekte um, die direkt als KI-Whitelist genutzt werden.
   → Die KI mappt SuSa-Konten auf genau diese Pool-Positionsnamen.

2. assign_dummy_ids()
   Sucht für jedes gemappte Konto den ersten freien Dummy-OID der zugewiesenen
   Pool-Überposition. Da die KI jetzt mit denselben Namen arbeitet, ist der
   Lookup immer exakt (kein Fuzzy, kein Fallback nötig).

Pool-Quelle: Reiter 'Mapping' in Dummykonten_Zuordnung_hart.xlsx
    Spalten: Überposition | Dummy Konto Name | OID | TargetDimensionID
"""
from __future__ import annotations

import logging
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
import pandas as pd

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Pool laden
# ---------------------------------------------------------------------------

def load_dummy_pool(dummy_xlsx: str | Path) -> Dict[str, List[dict]]:
    """
    Lädt Reiter 'Mapping' → Dict: normalize(überposition) → [{name, oid, dimension}, …]
    Reihenfolge bleibt erhalten (First-Come-First-Served).
    """
    dummy_xlsx = Path(dummy_xlsx)
    wb = openpyxl.load_workbook(str(dummy_xlsx), read_only=True, data_only=True)
    if "Mapping" not in wb.sheetnames:
        raise ValueError(f"Kein Reiter 'Mapping' in {dummy_xlsx}")

    ws = wb["Mapping"]
    pool: Dict[str, List[dict]] = defaultdict(list)

    header_skipped = False
    for row in ws.iter_rows(values_only=True):
        if not header_skipped:
            header_skipped = True
            continue
        ueberpos, dummy_name, oid, dimension = row[0], row[1], row[2], row[3]
        if not ueberpos or not dummy_name:
            continue
        key = _normalize(str(ueberpos))
        pool[key].append({
            "name":         dummy_name,
            "oid":          int(oid) if oid is not None else None,
            "dimension":    dimension,
            "ueberpos_raw": ueberpos,
        })

    wb.close()
    total = sum(len(v) for v in pool.values())
    logger.info("Dummy-Pool geladen: %d Überpositionen, %d Dummy-Konten", len(pool), total)
    return pool


# ---------------------------------------------------------------------------
# Pool-Überpositionen als TargetPosition-Objekte (für KI-Whitelist)
# ---------------------------------------------------------------------------

def pool_positions_as_targets(dummy_pool: Dict[str, List[dict]]):
    """
    Erstellt für jede eindeutige Überposition im Pool ein TargetPosition-Objekt.
    Die KI bekommt diese als Whitelist und mappt SuSa-Konten direkt auf diese Namen.

    Der target_id ist der normalisierte Positionsname (= dict-key im Pool),
    target_name ist der originale Name aus der ersten Pool-Zeile dieser Gruppe.
    target_class wird aus üblichen Konventionen abgeleitet (AKTIVA/PASSIVA/AUFWAND/ERTRAG).
    """
    from src.targets import TargetPosition

    positions = []
    for norm_key, entries in dummy_pool.items():
        raw_name = entries[0]["ueberpos_raw"]
        dimension = entries[0]["dimension"] or ""

        # Einfache Klassen-Heuristik an Hand der Dimension
        if "bilanz" in dimension.lower():
            target_class = "AKTIVA"  # wird von KI eh ignoriert bei direktem Namens-Match
        else:
            target_class = "ERTRAG"

        positions.append(TargetPosition(
            target_id=norm_key,          # normalisierter Key = späterer Lookup-Key
            target_name=raw_name,        # lesbarer Name für die KI
            target_class=target_class,
            is_leaf=True,
            level=1,
            sheet="DummyPool",
        ))

    logger.info("Pool-Positionen als Targets: %d", len(positions))
    return positions


# ---------------------------------------------------------------------------
# Dummy-OID Zuordnung
# ---------------------------------------------------------------------------

def assign_dummy_ids(
    mapped_df: pd.DataFrame,
    dummy_pool: Dict[str, List[dict]],
    target_name_col: str = "target_overpos_name",
) -> pd.DataFrame:
    """
    Ergänzt mapped_df um Spalten: dummy_name, dummy_oid, dummy_dimension

    Da die KI jetzt direkt auf Pool-Positionsnamen gemappt hat, ist
    der Lookup immer exakt (normalisierter Schlüssel → Pool-Eintrag).
    Keine Dummy-Konten verfügbar → Felder leer, Warning ins Log.
    """
    df = mapped_df.copy()
    pointer: Dict[str, int] = defaultdict(int)
    dummy_names, dummy_oids, dummy_dimensions = [], [], []

    for _, row in df.iterrows():
        if row.get("row_type") != "ACCOUNT":
            dummy_names.append("")
            dummy_oids.append(None)
            dummy_dimensions.append("")
            continue

        overpos_name = str(row.get(target_name_col, "") or "").strip()
        if not overpos_name or overpos_name in ("UNMAPPED", "nan", ""):
            dummy_names.append("")
            dummy_oids.append(None)
            dummy_dimensions.append("")
            continue

        # target_overpos_id enthält den normalisierten Pool-Key (gesetzt von map_accounts)
        overpos_id = str(row.get("target_overpos_id", "") or "").strip()
        key = overpos_id if overpos_id and overpos_id != "UNMAPPED" else _normalize(overpos_name)

        pool_list = dummy_pool.get(key, [])
        idx = pointer[key]

        if not pool_list:
            logger.warning("Keine Pool-Position für KI-Target '%s' (key='%s')", overpos_name, key)
            dummy_names.append("")
            dummy_oids.append(None)
            dummy_dimensions.append("")
        elif idx >= len(pool_list):
            logger.warning("Dummy-Pool erschöpft für '%s' (%d/%d belegt)",
                           overpos_name, idx, len(pool_list))
            dummy_names.append("")
            dummy_oids.append(None)
            dummy_dimensions.append("")
        else:
            dummy = pool_list[idx]
            dummy_names.append(dummy["name"])
            dummy_oids.append(dummy["oid"])
            dummy_dimensions.append(dummy["dimension"])
            pointer[key] += 1

    df["dummy_name"]      = dummy_names
    df["dummy_oid"]       = dummy_oids
    df["dummy_dimension"] = dummy_dimensions
    return df


# ---------------------------------------------------------------------------
# Lucanet-Export
# ---------------------------------------------------------------------------

_LUCANET_HEADER = [
    "SourceName", "TargetName", "TargetElementID", "TargetDimensionID",
    "Type", "DefaultCurrency", "DecimalDigits",
    "FirstPeriodOfFiscalYear", "StartMonth", "EndMonth", "AccountingAreaID",
]


def build_lucanet_df(mapped_df: pd.DataFrame) -> pd.DataFrame:
    """Baut den Lucanet-Import-DataFrame aus dem angereicherten mapped_df."""
    accounts = mapped_df[
        (mapped_df["row_type"] == "ACCOUNT") &
        (mapped_df["dummy_oid"].notna()) &
        (mapped_df["dummy_oid"] != "")
    ].copy()

    rows = []
    for _, row in accounts.iterrows():
        nr   = str(row.get("konto_nr",   "") or "").strip()
        name = str(row.get("konto_name", "") or "").strip()
        source_name = f"{nr} {name}".strip() if nr else name

        rows.append({
            "SourceName":              source_name,
            "TargetName":              row["dummy_name"],
            "TargetElementID":         int(row["dummy_oid"]),
            "TargetDimensionID":       row["dummy_dimension"],
            "Type":                    "Account",
            "DefaultCurrency":         "",
            "DecimalDigits":           0,
            "FirstPeriodOfFiscalYear": 0,
            "StartMonth":              "",
            "EndMonth":                "",
            "AccountingAreaID":        "",
        })

    return pd.DataFrame(rows, columns=_LUCANET_HEADER)


def save_lucanet_xlsx(lucanet_df: pd.DataFrame, out_path: str | Path) -> None:
    """Speichert den DataFrame als xlsx im Lucanet-Import-Format."""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(str(out_path), engine="openpyxl") as writer:
        lucanet_df.to_excel(writer, index=False, sheet_name="Lucanet Mapping")
        ws = writer.sheets["Lucanet Mapping"]
        for col_cells in ws.columns:
            max_len = max(len(str(c.value) if c.value is not None else "") for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)

    logger.info("Lucanet-Export gespeichert: %s (%d Zeilen)", out_path, len(lucanet_df))


# ---------------------------------------------------------------------------
# Hilfsfunktion
# ---------------------------------------------------------------------------

def _normalize(text: str) -> str:
    """Normalisiert für robusten Schlüsselvergleich."""
    return " ".join(text.strip().split()).lower()
