"""
generate_lucanet_mapping.py
----------------------------
Liest den KI-Mapping-Output (Mapping_Blatt_X.xlsx) und ordnet jedem
Quellkonto das erste freie Dummy-Konto der zugehörigen Lucanet-Überposition zu.

Ausgabeformat entspricht dem Reiter "KI_OUTPUT_Beispiel" aus
Dummykonten_Zuordnung_hart.xlsx:
    A: SourceName
    B: TargetName       (Dummy Konto Name)
    C: TargetElementID  (OID)
    D: TargetDimensionID (Bilanz / GuV)
    E: Type             = "Account"
    F: DefaultCurrency  = ""
    G: DecimalDigits    = 0
    H: FirstPeriodOfFiscalYear = 0
    I: StartMonth       = ""
    J: EndMonth         = ""
    K: AccountingAreaID = ""

Verwendung:
    python src/generate_lucanet_mapping.py \
        --ki-output output/sheet_5/Mapping_Blatt_5.xlsx \
        --dummy-pool Examples/Dummykonten_Zuordnung_hart.xlsx \
        --output output/sheet_5/Lucanet_Mapping_Blatt_5.xlsx
"""

import argparse
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl import Workbook


# ──────────────────────────────────────────────────────────────────────────────
# Hilfsfunktionen
# ──────────────────────────────────────────────────────────────────────────────

def normalize(text: str) -> str:
    """Normalisiert einen String für robuste Vergleiche."""
    if not isinstance(text, str):
        return ""
    return " ".join(text.strip().split()).lower()


def load_dummy_pool(dummy_xlsx: Path) -> dict[str, list[dict]]:
    """
    Lädt den Reiter 'Mapping' aus der Dummykonten-Datei.

    Gibt ein Dict zurück:
        normalized_überposition → [ {name, oid, dimension}, ... ]
    Die Liste enthält die Dummy-Konten in ihrer ursprünglichen Reihenfolge,
    damit wir immer das erste freie vergeben können.
    """
    wb = openpyxl.load_workbook(str(dummy_xlsx), read_only=True, data_only=True)
    if "Mapping" not in wb.sheetnames:
        raise ValueError(f"Kein Reiter 'Mapping' in {dummy_xlsx}")

    ws = wb["Mapping"]
    pool: dict[str, list[dict]] = defaultdict(list)

    header_skipped = False
    for row in ws.iter_rows(values_only=True):
        if not header_skipped:
            header_skipped = True
            continue  # Kopfzeile überspringen

        ueberpos, dummy_name, oid, dimension = row[0], row[1], row[2], row[3]
        if not ueberpos or not dummy_name:
            continue

        key = normalize(str(ueberpos))
        pool[key].append(
            {
                "name": dummy_name,
                "oid": int(oid) if oid is not None else None,
                "dimension": dimension,
                "ueberpos_raw": ueberpos,
            }
        )

    wb.close()
    return pool


def load_ki_output(ki_xlsx: Path) -> list[dict]:
    """
    Lädt den KI-Mapping-Output.

    Erwartet Spalten (Header in Zeile 1):
        Konto-Nr | Konto-Name | Kontonr & Bezeichnung |
        Unsere Zuordnung (LucaNet) | Target ID | Confidence | Begründung

    Gibt eine Liste von Dicts zurück.
    """
    wb = openpyxl.load_workbook(str(ki_xlsx), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Keine Daten in {ki_xlsx}")

    header = [str(h).strip() if h is not None else "" for h in rows[0]]

    # Spaltenindizes ermitteln – case-insensitive
    header_lower = [h.lower() for h in header]

    def col(candidates: list[str]) -> int:
        for c in candidates:
            for i, h in enumerate(header_lower):
                if c.lower() in h:
                    return i
        raise ValueError(f"Keine Spalte gefunden für: {candidates}")

    idx_nr        = col(["konto-nr", "konto nr", "kontonr"])
    idx_name      = col(["konto-name", "kontoname", "konto name"])
    idx_zuordnung = col(["unsere zuordnung", "lucanet"])

    accounts = []
    for row in rows[1:]:
        if not row or all(v is None for v in row):
            continue
        nr        = str(row[idx_nr]).strip()   if row[idx_nr]        is not None else ""
        name      = str(row[idx_name]).strip() if row[idx_name]      is not None else ""
        zuordnung = str(row[idx_zuordnung]).strip() if row[idx_zuordnung] is not None else ""

        if not zuordnung or zuordnung.lower() in ("none", ""):
            continue  # Ohne Zuordnung nicht verwendbar

        # SourceName = "Konto-Nr Konto-Name"
        source_name = f"{nr} {name}".strip() if nr else name

        accounts.append(
            {
                "source_name": source_name,
                "zuordnung_raw": zuordnung,
                "zuordnung_key": normalize(zuordnung),
            }
        )

    wb.close()
    return accounts


def build_output(ki_accounts: list[dict], dummy_pool: dict[str, list[dict]]) -> list[dict]:
    """
    Ordnet jedem KI-Konto das nächste freie Dummy-Konto zu.
    Gibt eine Liste von Ausgabe-Dicts zurück.
    """
    # Zeiger: für jede Überposition merken wir uns, welches Dummy als nächstes frei ist
    pointer: dict[str, int] = defaultdict(int)
    out_rows = []
    warnings = []

    for acc in ki_accounts:
        key = acc["zuordnung_key"]

        if key not in dummy_pool:
            warnings.append(
                f"  ⚠ Keine Dummy-Konten für Überposition: '{acc['zuordnung_raw']}'"
            )
            continue

        idx = pointer[key]
        pool_for_pos = dummy_pool[key]

        if idx >= len(pool_for_pos):
            warnings.append(
                f"  ⚠ Dummy-Pool erschöpft für '{acc['zuordnung_raw']}' "
                f"(benötigt: {idx + 1}, vorhanden: {len(pool_for_pos)})"
            )
            continue

        dummy = pool_for_pos[idx]
        pointer[key] += 1

        out_rows.append(
            {
                "SourceName":        acc["source_name"],
                "TargetName":        dummy["name"],
                "TargetElementID":   dummy["oid"],
                "TargetDimensionID": dummy["dimension"],
                # Feste Spalten E–K
                "Type":              "Account",
                "DefaultCurrency":   "",
                "DecimalDigits":     0,
                "FirstPeriodOfFiscalYear": 0,
                "StartMonth":        "",
                "EndMonth":          "",
                "AccountingAreaID":  "",
            }
        )

    if warnings:
        print("\n".join(warnings))

    return out_rows


def write_output(out_rows: list[dict], output_path: Path) -> None:
    """Schreibt das Ergebnis als xlsx im KI_OUTPUT_Beispiel-Format."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Lucanet Mapping"

    headers = [
        "SourceName",
        "TargetName",
        "TargetElementID",
        "TargetDimensionID",
        "Type",
        "DefaultCurrency",
        "DecimalDigits",
        "FirstPeriodOfFiscalYear",
        "StartMonth",
        "EndMonth",
        "AccountingAreaID",
    ]
    ws.append(headers)

    for row in out_rows:
        ws.append([row[h] for h in headers])

    # Spaltenbreiten anpassen
    for col_cells in ws.columns:
        max_len = max(len(str(c.value) if c.value is not None else "") for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"✅ Gespeichert: {output_path}  ({len(out_rows)} Zeilen)")


# ──────────────────────────────────────────────────────────────────────────────
# Hauptprogramm
# ──────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Generiert Lucanet-Mapping aus KI-Output")
    parser.add_argument(
        "--ki-output",
        required=True,
        help="Pfad zur KI-Output-Datei (z. B. output/sheet_5/Mapping_Blatt_5.xlsx)",
    )
    parser.add_argument(
        "--dummy-pool",
        default="Examples/Dummykonten_Zuordnung_hart.xlsx",
        help="Pfad zur Dummy-Pool-Datei (Standard: Examples/Dummykonten_Zuordnung_hart.xlsx)",
    )
    parser.add_argument(
        "--output",
        help="Ausgabepfad (Standard: gleicher Ordner wie --ki-output, Datei: Lucanet_<ki-filename>)",
    )
    args = parser.parse_args()

    ki_path   = Path(args.ki_output)
    pool_path = Path(args.dummy_pool)

    if not ki_path.exists():
        print(f"Fehler: KI-Output nicht gefunden: {ki_path}", file=sys.stderr)
        sys.exit(1)
    if not pool_path.exists():
        print(f"Fehler: Dummy-Pool nicht gefunden: {pool_path}", file=sys.stderr)
        sys.exit(1)

    if args.output:
        out_path = Path(args.output)
    else:
        out_path = ki_path.parent / f"Lucanet_{ki_path.name}"

    print(f"📂 KI-Output:   {ki_path}")
    print(f"📂 Dummy-Pool:  {pool_path}")
    print(f"📂 Zieldatei:   {out_path}")
    print()

    print("Lade Dummy-Pool …")
    dummy_pool = load_dummy_pool(pool_path)
    total_dummies = sum(len(v) for v in dummy_pool.values())
    print(f"  → {len(dummy_pool)} Überpositionen, {total_dummies} Dummy-Konten geladen")

    print("Lade KI-Output …")
    ki_accounts = load_ki_output(ki_path)
    print(f"  → {len(ki_accounts)} Konten mit Zuordnung geladen")

    print("Erstelle Mapping …")
    out_rows = build_output(ki_accounts, dummy_pool)
    print(f"  → {len(out_rows)} Konten erfolgreich zugeordnet")

    write_output(out_rows, out_path)


if __name__ == "__main__":
    main()
